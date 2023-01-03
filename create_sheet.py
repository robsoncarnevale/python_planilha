import openpyxl

#criar uma planilha(book)
book = openpyxl.Workbook()
# comando para visualizar as paginas
print(book.sheetnames)
#criando uma pagina
book.create_sheet('Vendas')
#selecionando uma pagina
vendas_page = book['Vendas']
#adicionando o cabeçlho 
vendas_page.append(['Animais', 'Qtda', 'Valor'])
#adicionando valores na pagina
vendas_page.append(['Cobra', 1, 'R$3.000,00'])
vendas_page.append(['Macaco Prego', 1, 'R$5.000,00'])
vendas_page.append(['Arara Vermelha', 1, 'R$10.000,00'])
vendas_page.append(['Papagaio', 1, 'R$3.500,00'])
vendas_page.append(['Lagarto', 1, 'R$2.000,00'])
#Salvar a Planilha
book.save('Planilha vendas de animais.xlsx')
